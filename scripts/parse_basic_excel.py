#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict

import openpyxl


def dump_workbook(path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=False)
    out: Dict[str, Any] = {"source": str(path), "sheets": {}}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        grid: list[list[Any]] = []
        for r in range(1, max_row + 1):
            row_vals: list[Any] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                value = cell.value
                if cell.data_type == "f":
                    value = f"={value}"
                row_vals.append(value)
            grid.append(row_vals)
        out["sheets"][sheet] = {
            "max_row": max_row,
            "max_col": max_col,
            "grid": grid,
        }
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse vancomycin Excel into JSON")
    parser.add_argument("excel_path", type=str, help="Path to XLSX file")
    parser.add_argument(
        "--out",
        type=str,
        default="data/parsed/basic_workbook.json",
        help="Output JSON path",
    )
    args = parser.parse_args()
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    data = dump_workbook(Path(args.excel_path))
    out_path.write_text(json.dumps(data, indent=2))
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
